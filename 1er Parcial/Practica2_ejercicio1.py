import csv
import heapq
import os
from collections import deque
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def load_maze(filename):
    with open(filename, 'r', newline='', encoding='utf-8-sig') as file:
        reader = csv.reader(file)
        maze = []
        for row in reader:
            cleaned_row = [cell.strip().replace('"', '').replace("'", "") for cell in row]
            maze.append(cleaned_row)
    return maze

def find_start_end(maze):
    start = None
    end = None
    for i, row in enumerate(maze):
        for j, cell in enumerate(row):
            if cell == 'S':
                start = (i, j)
            elif cell == 'E':
                end = (i, j)
    
    if start is None:
        raise ValueError("No se encontró el punto de inicio 'S'")
    if end is None:
        raise ValueError("No se encontró el punto final 'E'")
    
    return start, end

def heuristic(a, b):
    return abs(a[0] - b[0]) + abs(a[1] - b[1])

def a_star(maze, start, end):
    neighbors = [(0, 1), (1, 0), (0, -1), (-1, 0)]
    open_list = []
    heapq.heappush(open_list, (0, start))
    g_score = {start: 0}
    f_score = {start: heuristic(start, end)}
    came_from = {}
    closed_set = set()

    while open_list:
        _, current = heapq.heappop(open_list)
        if current == end:
            path = []
            while current in came_from:
                path.append(current)
                current = came_from[current]
            path.append(start)
            path.reverse()
            return path

        closed_set.add(current)
        for dx, dy in neighbors:
            neighbor = (current[0] + dx, current[1] + dy)
            
            if not (0 <= neighbor[0] < len(maze) and 0 <= neighbor[1] < len(maze[0])):
                continue
            
            if maze[neighbor[0]][neighbor[1]] == '1':
                continue
                
            tentative_g = g_score[current] + 1
            
            if neighbor in closed_set and tentative_g >= g_score.get(neighbor, float('inf')):
                continue
                
            if tentative_g < g_score.get(neighbor, float('inf')):
                came_from[neighbor] = current
                g_score[neighbor] = tentative_g
                f_score[neighbor] = tentative_g + heuristic(neighbor, end)
                heapq.heappush(open_list, (f_score[neighbor], neighbor))
    
    return None

def dfs(maze, start, end):
    stack = [(start, [start])]
    visited = set()
    
    while stack:
        (current, path) = stack.pop()
        if current in visited:
            continue
            
        visited.add(current)
        
        if current == end:
            return path
            
        for dx, dy in [(0, 1), (1, 0), (0, -1), (-1, 0)]:
            neighbor = (current[0] + dx, current[1] + dy)
            
            if not (0 <= neighbor[0] < len(maze) and 0 <= neighbor[1] < len(maze[0])):
                continue
                
            if maze[neighbor[0]][neighbor[1]] == '1':
                continue
                
            if neighbor not in visited:
                stack.append((neighbor, path + [neighbor]))
    
    return None

def bfs(maze, start, end):
    queue = deque([(start, [start])])
    visited = set()
    
    while queue:
        (current, path) = queue.popleft()
        if current in visited:
            continue
            
        visited.add(current)
        
        if current == end:
            return path
            
        for dx, dy in [(0, 1), (1, 0), (0, -1), (-1, 0)]:
            neighbor = (current[0] + dx, current[1] + dy)
            
            if not (0 <= neighbor[0] < len(maze) and 0 <= neighbor[1] < len(maze[0])):
                continue
                
            if maze[neighbor[0]][neighbor[1]] == '1':
                continue
                
            if neighbor not in visited:
                queue.append((neighbor, path + [neighbor]))
    
    return None

def create_excel_with_colored_solutions(maze, a_star_path, dfs_path, bfs_path, filename):
    # Crear un nuevo libro de trabajo de Excel
    wb = Workbook()
    
    # Definir colores
    start_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde para inicio
    end_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Rojo para fin
    path_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")   # Amarillo para camino
    obstacle_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Negro para obstáculos
    free_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")    # Blanco para celdas libres
    
    # Crear hojas para cada algoritmo
    algorithms = {
        "A_Star": a_star_path,
        "DFS": dfs_path,
        "BFS": bfs_path
    }
    
    for algo_name, path in algorithms.items():
        if algo_name == "A_Star":
            ws = wb.active
            ws.title = algo_name
        else:
            ws = wb.create_sheet(algo_name)
        
        # Escribir el laberinto en la hoja
        for i, row in enumerate(maze):
            for j, cell in enumerate(row):
                ws.cell(row=i+1, column=j+1, value=cell)
                
                # Aplicar colores según el tipo de celda
                if cell == 'S':
                    ws.cell(i+1, j+1).fill = start_fill
                elif cell == 'E':
                    ws.cell(i+1, j+1).fill = end_fill
                elif cell == '1':
                    ws.cell(i+1, j+1).fill = obstacle_fill
                else:
                    ws.cell(i+1, j+1).fill = free_fill
        
        # Resaltar el camino encontrado
        if path:
            for i, j in path:
                if maze[i][j] not in ['S', 'E']:
                    ws.cell(row=i+1, column=j+1).value = '*'
                    ws.cell(row=i+1, column=j+1).fill = path_fill
    
    # Ajustar el ancho de las columnas para mejor visualización
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 10)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar el archivo Excel
    wb.save(filename)
    print(f"Archivo Excel con soluciones guardado en: {os.path.abspath(filename)}")

def main():
    try:
        input_filename = 'C:/Users/Rodri/OneDrive/Documentos/GitHub/IA/1er Parcial/laberinto.csv'
        output_filename = 'laberinto_soluciones.xlsx'
        
        maze = load_maze(input_filename)        
        start, end = find_start_end(maze)
        
        print("Ejecutando A*...")
        a_star_path = a_star(maze, start, end)
        
        print("Ejecutando DFS...")
        dfs_path = dfs(maze, start, end)
        
        print("Ejecutando BFS...")
        bfs_path = bfs(maze, start, end)
        
        if a_star_path and dfs_path and bfs_path:
            print(f"A* encontró un camino de {len(a_star_path)} pasos")
            print(f"DFS encontró un camino de {len(dfs_path)} pasos")
            print(f"BFS encontró un camino de {len(bfs_path)} pasos")
            
            # Crear archivo Excel con las soluciones
            create_excel_with_colored_solutions(maze, a_star_path, dfs_path, bfs_path, output_filename)
            
            print("Proceso completado. Revise el archivo Excel generado.")
        else:
            print("Al menos uno de los algoritmos no encontró un camino válido.")
            
    except FileNotFoundError:
        print(f"Error: No se pudo encontrar el archivo '{input_filename}'")
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == '__main__':
    main()